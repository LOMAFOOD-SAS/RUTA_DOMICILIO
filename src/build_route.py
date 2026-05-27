#!/usr/bin/env python3
"""LOMAFOOD daily route builder.

Lee un Excel con la lista de clientes del día y genera:
  * un HTML con la ruta optimizada
  * un JSON sidecar con metadata (destinatario, distancia, URL de Google Maps, ...)

En producción se ejecuta diariamente a las 4:30 PM desde Cowork (Claude desktop)
y un flujo de Power Automate envía el HTML como correo.

Para correrlo con el template incluido:
    python3 src/build_route.py --xlsx templates/AA_RUTA_DEL_DIA_TEMPLATE.xlsx \
                               --out  examples/

Para correrlo en producción (Cowork) sin argumentos: detecta el Excel real en
`/sessions/*/mnt/(08) Ruta/AA_RUTA DEL DIA.xlsx` y escribe en la misma carpeta.
"""
import argparse
import glob
import json
import math
import os
import re
import sys
import warnings
from datetime import datetime, timedelta
from itertools import permutations
from urllib.parse import quote

warnings.filterwarnings("ignore")
import openpyxl

# --- Defaults (parametrizables por flag) ---
DEFAULT_ORIGIN_ADDR = "Calle 77 #58-35, Bogota, Colombia"
DEFAULT_ORIGIN_COORDS = (4.6685, -74.0589)  # haversine seed only — no network calls
DEFAULT_MAX_KM = 70.0
DEFAULT_RECIPIENT = "pedidos@lomafood.com"


def parse_args():
    p = argparse.ArgumentParser(description="Genera la ruta diaria optimizada.")
    p.add_argument("--xlsx", help="Ruta al .xlsx fuente. Si se omite, se busca con glob en Cowork.")
    p.add_argument("--out", help="Carpeta donde escribir los outputs. Default: misma del xlsx.")
    p.add_argument("--origin", default=DEFAULT_ORIGIN_ADDR, help="Dirección de la bodega.")
    p.add_argument("--origin-coords", default=f"{DEFAULT_ORIGIN_COORDS[0]},{DEFAULT_ORIGIN_COORDS[1]}",
                   help="Coordenadas de la bodega, formato 'lat,lng'.")
    p.add_argument("--max-km", type=float, default=DEFAULT_MAX_KM, help="Distancia máxima sin warning.")
    p.add_argument("--recipient", default=DEFAULT_RECIPIENT, help="Email destino en el sidecar JSON.")
    p.add_argument("--date", help="Fecha objetivo ISO (YYYY-MM-DD). Default: mañana.")
    return p.parse_args()


def resolve_xlsx_path(arg_path):
    if arg_path:
        return arg_path
    candidates = glob.glob("/sessions/*/mnt/(08) Ruta/AA_RUTA DEL DIA.xlsx")
    if not candidates:
        print("ERROR: --xlsx no provisto y no encuentro el Excel real en Cowork.", file=sys.stderr)
        sys.exit(2)
    return candidates[0]


def hav(a, b):
    R = 6371.0
    p1, p2 = math.radians(a[0]), math.radians(b[0])
    dp = math.radians(b[0] - a[0])
    dl = math.radians(b[1] - a[1])
    h = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * R * math.asin(math.sqrt(h))


def total_distance(order, clients, origin_coords):
    pts = [origin_coords] + [clients[i]["coords"] for i in order] + [origin_coords]
    return sum(hav(pts[i], pts[i + 1]) for i in range(len(pts) - 1))


def cons_ok(order, active):
    n = len(order)
    for i, idx in enumerate(order):
        k = active[idx]
        if k == "first_only" and i != 0:
            return False
        if k == "first_or_second" and i not in (0, 1):
            return False
        if k == "last_or_prev" and i not in (n - 1, n - 2):
            return False
    return True


def classify_rule(text):
    rl = (text or "").lower()
    if "siempre primero" in rl:
        return "first_only"
    if "primero" in rl or "segundo" in rl or "prioritario" in rl:
        return "first_or_second"
    if "ultim" in rl or "penultim" in rl:
        return "last_or_prev"
    return None


def optimize(clients, constraints, origin_coords):
    n = len(clients)
    indices = list(range(n))
    if n <= 10:
        best = None
        best_d = float("inf")
        for perm in permutations(indices):
            if not cons_ok(perm, constraints):
                continue
            d = total_distance(perm, clients, origin_coords)
            if d < best_d:
                best_d = d
                best = list(perm)
        return best, best_d

    firsts = [i for i in indices if constraints[i] in ("first_only", "first_or_second")]
    lasts = [i for i in indices if constraints[i] == "last_or_prev"]
    free = [i for i in indices if constraints[i] not in ("first_only", "first_or_second", "last_or_prev")]

    if firsts:
        start = firsts.pop(0)
    else:
        pool = free + lasts
        start = min(pool, key=lambda i: hav(origin_coords, clients[i]["coords"]))
        (free if start in free else lasts).remove(start)

    order = [start]
    pool = free + firsts
    cur = clients[start]["coords"]
    while pool:
        nxt = min(pool, key=lambda i: hav(cur, clients[i]["coords"]))
        order.append(nxt)
        pool.remove(nxt)
        cur = clients[nxt]["coords"]
    for l in lasts:
        order.append(l)

    improved = True
    while improved:
        improved = False
        for i in range(len(order) - 1):
            for j in range(i + 1, len(order)):
                new = order[:i] + order[i:j + 1][::-1] + order[j + 1:]
                if not cons_ok(new, constraints):
                    continue
                if total_distance(new, clients, origin_coords) < total_distance(order, clients, origin_coords):
                    order = new
                    improved = True

    if cons_ok(order, constraints):
        return order, total_distance(order, clients, origin_coords)
    return None, None


def safe_write(path, content):
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        return path
    except (OSError, PermissionError, FileNotFoundError) as e:
        suffix = datetime.now().strftime("%H%M%S")
        base, ext = os.path.splitext(path)
        new_path = f"{base}-{suffix}{ext}"
        print(f"WARN: ghost en {path} ({e.__class__.__name__}); usando {new_path}", file=sys.stderr)
        with open(new_path, "w", encoding="utf-8") as f:
            f.write(content)
        return new_path


def main():
    args = parse_args()
    xlsx_path = resolve_xlsx_path(args.xlsx)
    out_dir = args.out or os.path.dirname(xlsx_path) or "."
    os.makedirs(out_dir, exist_ok=True)

    try:
        olat, olng = [float(x) for x in args.origin_coords.split(",")]
        origin_coords = (olat, olng)
    except Exception:
        print(f"ERROR: --origin-coords mal formada: {args.origin_coords}", file=sys.stderr)
        sys.exit(6)
    origin_addr = args.origin

    if args.date:
        target = datetime.strptime(args.date, "%Y-%m-%d")
    else:
        target = datetime.now() + timedelta(days=1)
    iso_date = target.strftime("%Y-%m-%d")
    display_date = target.strftime("%d/%m/%Y")

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"ERROR: No se pudo abrir {xlsx_path} ({type(e).__name__}: {e}).", file=sys.stderr)
        print("Para arreglarlo: abrir en Excel y 'Guardar como...' sobre el mismo nombre.", file=sys.stderr)
        sys.exit(3)

    master_sheet = next((wb[n] for n in wb.sheetnames if n.strip().upper() == "DATA VALIDATION"), None)
    if master_sheet is None:
        print("ERROR: falta la hoja 'DATA VALIDATION'.", file=sys.stderr)
        sys.exit(4)

    master = {}
    for row in master_sheet.iter_rows(min_row=2, min_col=2, max_col=6, values_only=True):
        cliente, direccion, latlng, ciudad, reglas = row
        if not cliente or not direccion or not latlng:
            continue
        try:
            parts = str(latlng).split(",")
            lat, lng = float(parts[0].strip()), float(parts[1].strip())
        except Exception:
            continue
        master[str(cliente).strip().lower()] = {
            "cliente": str(cliente).strip(),
            "direccion": str(direccion).strip(),
            "coords": (lat, lng),
            "ciudad": (str(ciudad).strip() if ciudad else ""),
            "reglas": (str(reglas).strip() if reglas else ""),
        }

    day_sheet = next((wb[n] for n in wb.sheetnames if n.strip().lower() == "sheet1"), wb.worksheets[0])

    today_clients = []
    warnings_list = []
    for row in day_sheet.iter_rows(min_row=3, max_col=4, values_only=True):
        cliente, direccion, ciudad, rule = row
        if not cliente:
            continue
        if direccion in (None, "", 0, "0"):
            continue
        key = str(cliente).strip().lower()
        m = master.get(key)
        if m is None:
            warnings_list.append(f"MISSING-FROM-MASTER: '{cliente}' no está en DATA VALIDATION; omitido.")
            continue
        sheet_rule = ""
        if rule not in (None, "", 0, "0"):
            sheet_rule = str(rule).strip()
        effective_rule = sheet_rule if sheet_rule else m["reglas"]
        today_clients.append({
            "cliente": m["cliente"],
            "direccion": m["direccion"],
            "ciudad": m["ciudad"],
            "coords": m["coords"],
            "rule": effective_rule,
            "rule_kind": classify_rule(effective_rule),
        })

    if not today_clients:
        skip_path = os.path.join(out_dir, f"ruta-{iso_date}.SKIP")
        try:
            with open(skip_path, "w", encoding="utf-8") as f:
                f.write("Sin clientes hoy")
            print(f"Sin clientes. Escribí {skip_path}.")
        except Exception as e:
            print(f"Sin clientes y no pude escribir SKIP ({e}).")
        return 0

    # Demote duplicate "siempre primero"
    first_only_idxs = [i for i, c in enumerate(today_clients) if c["rule_kind"] == "first_only"]
    if len(first_only_idxs) > 1:
        for i in first_only_idxs[1:]:
            warnings_list.append(
                f"DEMOTE: '{today_clients[i]['cliente']}' tenía 'siempre primero' duplicado; "
                "degradado a first_or_second."
            )
            today_clients[i]["rule_kind"] = "first_or_second"

    constraints = {i: c["rule_kind"] for i, c in enumerate(today_clients)}
    relaxed_log = []

    best_order, best_dist = optimize(today_clients, constraints, origin_coords)
    if best_order is None:
        relaxed = {i: (None if k == "first_or_second" else k) for i, k in constraints.items()}
        if relaxed != constraints:
            relaxed_log.append("Relajado: first_or_second.")
            best_order, best_dist = optimize(today_clients, relaxed, origin_coords)
    if best_order is None:
        relaxed = {i: (None if k in ("first_only", "first_or_second") else k) for i, k in constraints.items()}
        relaxed_log.append("Relajado: first_only.")
        best_order, best_dist = optimize(today_clients, relaxed, origin_coords)
    if best_order is None:
        relaxed = {i: None for i in constraints}
        relaxed_log.append("Relajado: last_or_prev.")
        best_order, best_dist = optimize(today_clients, relaxed, origin_coords)
    if best_order is None:
        print("ERROR: no se pudo construir ruta válida.", file=sys.stderr)
        return 5

    ordered = [today_clients[i] for i in best_order]
    origin_enc = quote(origin_addr)
    waypoints_enc = "|".join(quote(c["direccion"]) for c in ordered)
    url = (
        f"https://www.google.com/maps/dir/?api=1"
        f"&origin={origin_enc}&destination={origin_enc}"
        f"&waypoints={waypoints_enc}&travelmode=driving"
    )
    total_km = round(best_dist, 1)

    pts = (
        [(origin_addr, origin_coords)]
        + [(c["cliente"], c["coords"]) for c in ordered]
        + [(origin_addr, origin_coords)]
    )
    seg = [(hav(pts[i][1], pts[i + 1][1]), pts[i][0], pts[i + 1][0]) for i in range(len(pts) - 1)]
    longest = max(seg, key=lambda x: x[0])
    exceeds = total_km > args.max_km

    items_html = "\n".join(f'  <li>{c["cliente"]} — {c["direccion"]}</li>' for c in ordered)
    warn_html = ""
    if exceeds:
        warn_html = (
            f'<p style="color:#c00;"><strong>⚠️ ADVERTENCIA:</strong> '
            f'La ruta excede {args.max_km:g} km. Considere dividir la ruta. '
            f'Segmento más largo: {longest[1]} → {longest[2]} ({round(longest[0], 1)} km).</p>'
        )

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Ruta del día - {display_date}</title></head>
<body style="font-family: Arial, sans-serif; color: #222;">
<p>Buenas tardes,</p>
<p>Esta es la ruta optimizada para mañana (<strong>{display_date}</strong>):</p>
<p><strong>Orden de visita:</strong></p>
<ol>
{items_html}
  <li>Regreso a bodega — {origin_addr}</li>
</ol>
<p><strong>Distancia total estimada:</strong> {total_km} km</p>
{warn_html}
<p><a href="{url}" style="display:inline-block;padding:10px 18px;background:#1a73e8;color:#fff;text-decoration:none;border-radius:4px;">Abrir en Google Maps</a></p>
<p style="margin-top:24px;color:#888;font-size:12px;">LOMAFOOD — Ruta generada automáticamente</p>
</body></html>
"""

    html_path = os.path.join(out_dir, f"ruta-{iso_date}.html")
    json_path = os.path.join(out_dir, f"ruta-{iso_date}.json")
    actual_html = safe_write(html_path, html)
    sidecar = {
        "to": args.recipient,
        "subject": f"Ruta del día - {display_date}",
        "distance_km": total_km,
        "stops": len(ordered),
        "exceeds_limit": exceeds,
        "google_maps_url": url,
        "html_file": os.path.basename(actual_html),
    }
    actual_json = safe_write(json_path, json.dumps(sidecar, ensure_ascii=False, indent=2))

    # Cleanup older than 14 days
    cutoff = (datetime.now() - timedelta(days=14)).date()
    pat = re.compile(r"^ruta-(\d{4})-(\d{2})-(\d{2})\.(html|json|SKIP)$")
    deleted = []
    for fn in os.listdir(out_dir):
        m = pat.match(fn)
        if not m:
            continue
        y, mo, d, _ = m.groups()
        try:
            fdate = datetime(int(y), int(mo), int(d)).date()
        except ValueError:
            continue
        if fdate < cutoff:
            try:
                os.remove(os.path.join(out_dir, fn))
                deleted.append(fn)
            except Exception as e:
                print(f"WARN: no borré {fn}: {e}", file=sys.stderr)

    print("=" * 60)
    print(f"Ruta generada para: {display_date} ({iso_date})")
    print(f"Stops: {len(ordered)}")
    print(f"Distancia total: {total_km} km")
    print(f"Excede {args.max_km:g} km: {exceeds}")
    if exceeds:
        print(f"Segmento más largo: {longest[1]} -> {longest[2]} ({round(longest[0], 1)} km)")
    if relaxed_log:
        print("Reglas relajadas:")
        for r in relaxed_log:
            print(f"  - {r}")
    if warnings_list:
        print("Avisos:")
        for w in warnings_list:
            print(f"  - {w}")
    if deleted:
        print(f"Archivos viejos borrados: {', '.join(deleted)}")
    print(f"HTML: {actual_html}")
    print(f"JSON: {actual_json}")
    print(f"URL : {url}")
    print("=" * 60)
    print("Orden:")
    for i, c in enumerate(ordered, 1):
        print(f"  {i}. {c['cliente']} [{c['rule_kind'] or '-'}]")
    return 0


if __name__ == "__main__":
    sys.exit(main())
